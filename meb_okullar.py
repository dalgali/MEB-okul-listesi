#!/usr/bin/env python3
"""MEB okul/kurum listesini çeken script.

Kaynak: https://www.meb.gov.tr/baglantilar/okullar/index.php
Sayfa, okul listesini DataTables ile `okullar_ajax.php` uç noktasından
JSON olarak alıyor. Bu script o uç noktayı sayfalayarak tüm Türkiye'yi
tek geçişte indirir ve CSV / JSON / XLSX / SQLite olarak yazar.

Çekilen alanlar:
    il, ilce, okul_adi, kurum_kodu, okul_turu, web_sitesi,
    adres, telefon, harita, cekim_tarihi

adres ve telefon bu "hızlı katman"da boş bırakılır; her okulun kendi
k12.tr sitesine ayrı ayrı girmek gerektiği için ileride ikinci bir
geçişle doldurulabilir. harita, kaynakta her okul için var olan konum
sayfasının adresidir.
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
import math
import sqlite3
import sys
import time
import warnings

warnings.filterwarnings("ignore")  # LibreSSL/urllib3 NotOpenSSLWarning'ini bastır

import requests  # noqa: E402

INDEX_URL = "https://www.meb.gov.tr/baglantilar/okullar/index.php"
AJAX_URL = "https://www.meb.gov.tr/baglantilar/okullar/okullar_ajax.php"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

COLUMNS = [
    "il",
    "ilce",
    "okul_adi",
    "kurum_kodu",
    "okul_turu",
    "web_sitesi",
    "adres",
    "telefon",
    "harita",
    "cekim_tarihi",
]

PAGE_SIZE = 1000
MAX_RETRY = 4

# --- Okul türü çıkarımı -----------------------------------------------------

# Türkçe büyük/küçük harf sorunlarını atlamak için ada bu normalizasyon
# uygulanır, sonra küçük harfli anahtar kelimeler aranır.
_TR_MAP = str.maketrans(
    {
        "İ": "i", "I": "ı", "Ş": "ş", "Ğ": "ğ",
        "Ü": "ü", "Ö": "ö", "Ç": "ç",
        "Î": "i", "î": "i", "Â": "a", "â": "a", "Û": "u", "û": "u",
    }
)

# Kaynaktaki tutarsız il yazımları (tr_title uygulandıktan sonra eşlenir)
_IL_ALIAS = {"Afyon": "Afyonkarahisar"}

_TR_CFLEX = str.maketrans({"Â": "A", "â": "a", "Î": "İ", "î": "i",
                           "Û": "U", "û": "u"})


def tr_lower(s: str) -> str:
    """Türkçe kurallarıyla küçük harf ('I'→'ı', 'İ'→'i')."""
    return s.replace("İ", "i").replace("I", "ı").translate(_TR_CFLEX).lower()


def tr_title(s: str) -> str:
    """Türkçe'ye uygun 'Başlık Düzeni': ANKARA→Ankara, IĞDIR→Iğdır,
    DİYARBAKIR→Diyarbakır, İZMİR→İzmir. Python'un .title()'ı bunları bozar."""
    _up = {"i": "İ", "ı": "I"}
    out = []
    cap = True
    for ch in " ".join(s.split()):
        low = tr_lower(ch)
        if ch.isalpha():
            out.append(_up.get(low, low.upper()) if cap else low)
            cap = False
        else:
            out.append(ch)
            cap = ch in " -/(.'’–"
    return "".join(out)


_FOLD = str.maketrans("çğıöşü", "cgiosu")


def _fold(s: str) -> str:
    """Türkçe karakter (ı i ğ ü ç ş ö) ve büyük/küçük harf duyarsız anahtar:
    Bağcılar ile Bagcılar aynı anahtara düşer."""
    return tr_lower(s).translate(_FOLD)


def canonicalize(rows: list[dict]) -> list[dict]:
    """Aynı il/ilçenin Türkçe karakter içermeyen yazılmış kopyalarını
    (kaynak hatası) en sık geçen doğru yazıma eşitler.
    Bağcılar/Bagcılar, Kadıköy/Kadiköy…"""
    from collections import Counter

    for field in ("il", "ilce"):
        variants: dict[str, Counter] = {}
        for r in rows:
            v = r[field]
            if v:
                variants.setdefault(_fold(v), Counter())[v] += 1
        canon = {k: c.most_common(1)[0][0] for k, c in variants.items()}
        for r in rows:
            if r[field]:
                r[field] = canon[_fold(r[field])]
    return rows


def _norm(s: str) -> str:
    # küçük harfe indir + çoklu boşlukları teke düşür (kaynakta "Mesleki  Eğitim
    # Merkezi" gibi çift boşluklu adlar var)
    return " ".join(s.translate(_TR_MAP).lower().split())


# (anahtar kelimeler, etiket) — sıra önemlidir, ilk eşleşen kazanır.
_TYPE_RULES = [
    (("bilim ve sanat merkezi", "bilsem"), "Bilim ve Sanat Merkezi"),
    (("rehberlik ve araştırma merkezi",), "Rehberlik ve Araştırma Merkezi"),
    (("halk eğitim",), "Halk Eğitimi Merkezi"),
    (("mesleki eğitim merkezi", "mesleki eğitimi merkezi", "meslekî eğitim"),
     "Mesleki Eğitim Merkezi"),
    (("hizmetiçi eğitim enstitüsü", "hizmet içi eğitim enstitüsü"),
     "Hizmet İçi Eğitim Enstitüsü"),
    (("olgunlaşma enstitüsü",), "Olgunlaşma Enstitüsü"),
    (("eğitim ve uygulama merkezi", "eğitim uygulama merkezi",
      "eğitim ve uygulama okulu", "eğitim uygulaması"),
     "Özel Eğitim Kurumu"),
    (("öğretmenevi", "öğretmen evi"), "Öğretmenevi"),
    (("anaokulu", "ana okulu", "okul öncesi"), "Anaokulu"),
    (("imam hatip ortaokulu", "imam-hatip ortaokulu"), "İmam Hatip Ortaokulu"),
    (("imam hatip lisesi", "imam-hatip lisesi", "anadolu imam hatip"),
     "Anadolu İmam Hatip Lisesi"),
    (("fen lisesi",), "Fen Lisesi"),
    (("sosyal bilimler lisesi",), "Sosyal Bilimler Lisesi"),
    (("güzel sanatlar lisesi",), "Güzel Sanatlar Lisesi"),
    (("spor lisesi",), "Spor Lisesi"),
    (("mesleki ve teknik", "meslek lisesi", "teknik lise",
      "kız meslek", "ticaret meslek", "sağlık meslek", "endüstri meslek"),
     "Mesleki ve Teknik Anadolu Lisesi"),
    (("çok programlı",), "Çok Programlı Anadolu Lisesi"),
    (("anadolu lisesi",), "Anadolu Lisesi"),
    (("açık öğretim lisesi", "açık lise", "akşam lisesi"), "Açık/Akşam Lisesi"),
    (("lisesi", "lise "), "Lise"),
    (("ortaokulu", "orta okulu", "ortaokul"), "Ortaokul"),
    (("ilkokulu", "ilk okulu", "ilkokul", "ilköğretim",
      "ılkokul", "ilkokolu", "illkokul", "ilkoklu"), "İlkokul"),
    (("özel eğitim",), "Özel Eğitim Kurumu"),
    (("yatılı bölge", "ybo"), "Yatılı Bölge Ortaokulu"),
    (("pansiyon",), "Pansiyon"),
    (("akşam sanat okulu", "sanat okulu"), "Akşam Sanat Okulu"),
    (("eğitim müşavirliği", "eğitim ataşeliği",
      "türkçe eğitim öğretim merkezi", "türkçe e.ö.mer", "türkç.e.ö.mer",
      "e.ö.merk"), "Yurt Dışı Teşkilatı"),
    (("genel müdürlüğü", "genel müdürlük", "bakanlık"),
     "Bakanlık Merkez Birimi"),
    (("milli eğitim müdürlüğü", "eğitim müdürlüğü"), "Milli Eğitim Müdürlüğü"),
]


def infer_type(ad: str) -> str:
    """Okul adından türü tahmin eder; eşleşme yoksa 'Belirsiz'."""
    n = _norm(ad or "")
    for keywords, label in _TYPE_RULES:
        for kw in keywords:
            if kw in n:
                return label
    return "Belirsiz"


# --- Satır ayrıştırma ------------------------------------------------------


def parse_row(raw: dict, cekim_tarihi: str) -> dict:
    """AJAX satırını (OKUL_ADI / HOST / YOL) çıktı sözlüğüne çevirir."""
    name = (raw.get("OKUL_ADI") or "").strip()
    bits = name.split(" - ", 2)
    if len(bits) == 3:
        il, ilce, okul_adi = bits[0].strip(), bits[1].strip(), bits[2].strip()
    elif len(bits) == 2:
        il, ilce, okul_adi = bits[0].strip(), bits[1].strip(), ""
    else:
        il, ilce, okul_adi = "", "", name

    # Kaynak il/ilçeyi TÜMÜ BÜYÜK verir; okunaklı yazıma çevir.
    il = tr_title(il) if il else ""
    ilce = tr_title(ilce) if ilce else ""
    il = _IL_ALIAS.get(il, il)

    yol = (raw.get("YOL") or "").strip()
    segs = [s for s in yol.split("/") if s]
    last = segs[-1] if segs else ""
    kurum_kodu = last if (last.isdigit() and len(last) == 6) else ""

    host = (raw.get("HOST") or "").strip()
    web_sitesi = f"https://{host}.meb.k12.tr" if host else ""
    harita = f"https://{host}.meb.k12.tr/tema/harita.php" if host else ""

    return {
        "il": il,
        "ilce": ilce,
        "okul_adi": okul_adi,
        "kurum_kodu": kurum_kodu,
        "okul_turu": infer_type(okul_adi or name),
        "web_sitesi": web_sitesi,
        "adres": "",
        "telefon": "",
        "harita": harita,
        "cekim_tarihi": cekim_tarihi,
    }


# --- Ağ ------------------------------------------------------------------


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": USER_AGENT,
            "X-Requested-With": "XMLHttpRequest",
            "Referer": INDEX_URL,
            "Accept": "application/json, text/javascript, */*; q=0.01",
        }
    )
    # Çerezleri ve oturumu ısıt; Referer olmadan uç nokta "Erişim yetkiniz
    # yok!" döndürüyor.
    r = s.get(INDEX_URL, timeout=30)
    r.raise_for_status()
    return s


def fetch_page(session: requests.Session, il: int, ilce: int,
               start: int, length: int) -> dict:
    data = {
        "draw": 1,
        "start": start,
        "length": length,
        "il": il,
        "ilce": ilce,
        "search[value]": "",
        "search[regex]": "false",
        "order[0][column]": 0,
        "order[0][dir]": "asc",
        "columns[0][data]": "OKUL_ADI",
    }
    last_err = None
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = session.post(AJAX_URL, data=data, timeout=45)
            r.raise_for_status()
            body = r.text.strip()
            if body.startswith("{"):
                return r.json()
            raise ValueError(f"beklenmeyen yanıt: {body[:120]!r}")
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            wait = 2 ** attempt
            print(f"  ! sayfa (start={start}) hata: {exc} — {wait}s sonra "
                  f"tekrar ({attempt}/{MAX_RETRY})", file=sys.stderr)
            time.sleep(wait)
    raise RuntimeError(f"sayfa start={start} {MAX_RETRY} denemede alınamadı: "
                       f"{last_err}")


def iter_rows(session: requests.Session, il: int = 0, ilce: int = 0,
              page_size: int = PAGE_SIZE, delay: float = 0.3):
    """Tüm satırları sayfalayarak üretir (ham AJAX sözlükleri)."""
    start = 0
    total = None
    pages_done = 0
    max_pages = 10_000
    while total is None or start < total:
        js = fetch_page(session, il, ilce, start, page_size)
        total = int(js.get("recordsTotal") or 0)
        batch = js.get("data") or []
        if not batch:
            break
        for raw in batch:
            yield raw, total
        start += len(batch)
        pages_done += 1
        done = min(start, total)
        pct = (done / total * 100) if total else 0
        print(f"  {done:>6}/{total} (%{pct:4.1f})")
        if pages_done > max_pages:
            print("  ! sayfa üst sınırına ulaşıldı, durduruluyor",
                  file=sys.stderr)
            break
        if start < total and delay:
            time.sleep(delay)


# --- Çıktı yazıcılar ----------------------------------------------------


def write_csv(rows: list[dict], path: str) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)


def write_json(rows: list[dict], path: str) -> None:
    # Geçerli JSON dizisi; satır başına bir kayıt (git diff'i için okunaklı).
    with open(path, "w", encoding="utf-8") as f:
        f.write("[\n")
        for i, r in enumerate(rows):
            f.write(json.dumps(r, ensure_ascii=False, separators=(",", ":")))
            f.write(",\n" if i < len(rows) - 1 else "\n")
        f.write("]\n")


def write_xlsx(rows: list[dict], path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  ! openpyxl kurulu değil, xlsx atlanıyor "
              "(pip install openpyxl)", file=sys.stderr)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "okullar"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
    ws.freeze_panes = "A2"
    wb.save(path)


def write_sqlite(rows: list[dict], path: str) -> None:
    import os

    if os.path.exists(path):
        os.remove(path)
    con = sqlite3.connect(path)
    try:
        cols_sql = ", ".join(f'"{c}" TEXT' for c in COLUMNS)
        con.execute(f"CREATE TABLE okullar ({cols_sql})")
        con.executemany(
            f"INSERT INTO okullar VALUES ({', '.join('?' for _ in COLUMNS)})",
            [tuple(r[c] for c in COLUMNS) for r in rows],
        )
        con.execute("CREATE INDEX ix_il ON okullar(il)")
        con.execute("CREATE INDEX ix_kurum ON okullar(kurum_kodu)")
        con.commit()
    finally:
        con.close()


WRITERS = {
    "csv": ("meb-okullar.csv", write_csv),
    "json": ("meb-okullar.json", write_json),
    "xlsx": ("meb-okullar.xlsx", write_xlsx),
    "sqlite": ("meb-okullar.sqlite", write_sqlite),
}


# --- Orkestrasyon -----------------------------------------------------------


def run(out_dir: str, formats: list[str], il: int, delay: float) -> int:
    import os

    os.makedirs(out_dir, exist_ok=True)
    cekim_tarihi = _dt.date.today().isoformat()

    print(f"MEB'e bağlanılıyor… (Referer oturumu ısıtılıyor)")
    session = make_session()

    scope = "tüm Türkiye" if il == 0 else f"il kodu {il}"
    print(f"Okullar çekiliyor — {scope}")

    seen = {}
    dup = 0
    expected = None
    for raw, total in iter_rows(session, il=il, delay=delay):
        expected = total
        row = parse_row(raw, cekim_tarihi)
        key = row["kurum_kodu"] or f"__{raw.get('HOST')}__{raw.get('OKUL_ADI')}"
        if key in seen:
            dup += 1
            continue
        seen[key] = row

    rows = canonicalize(list(seen.values()))
    rows.sort(key=lambda r: (r["il"], r["ilce"], r["okul_adi"]))
    print(f"\nToplam benzersiz kayıt: {len(rows)}  (yinelenen atlandı: {dup})")
    if expected is not None and abs(len(rows) + dup - expected) > 0:
        print(f"  ! uyarı: kaynak {expected} kayıt bildirdi, "
              f"{len(rows) + dup} işlendi", file=sys.stderr)

    missing_kod = sum(1 for r in rows if not r["kurum_kodu"])
    missing_web = sum(1 for r in rows if not r["web_sitesi"])
    belirsiz = sum(1 for r in rows if r["okul_turu"] == "Belirsiz")
    print(f"  kurum_kodu boş: {missing_kod}  |  web_sitesi boş: {missing_web}"
          f"  |  türü 'Belirsiz': {belirsiz}")

    for fmt in formats:
        fname, writer = WRITERS[fmt]
        path = os.path.join(out_dir, fname)
        writer(rows, path)
        if os.path.exists(path):
            kb = os.path.getsize(path) / 1024
            print(f"  yazıldı: {path}  ({kb:,.0f} KB)")

    print("Tamamlandı.")
    return 0


# --- Selftest (ağsız) ----------------------------------------------------


def selftest() -> int:
    t = "2026-08-31"

    # Türkçe başlık düzeni (Python .title()'ın bozduğu iller)
    assert tr_title("DİYARBAKIR") == "Diyarbakır", tr_title("DİYARBAKIR")
    assert tr_title("IĞDIR") == "Iğdır", tr_title("IĞDIR")
    assert tr_title("ŞANLIURFA") == "Şanlıurfa", tr_title("ŞANLIURFA")
    assert tr_title("KAHRAMANMARAŞ") == "Kahramanmaraş", tr_title("KAHRAMANMARAŞ")
    assert tr_title("İSTANBUL") == "İstanbul", tr_title("İSTANBUL")
    assert tr_title("AFYONKARAHİSAR") == "Afyonkarahisar", tr_title("AFYONKARAHİSAR")
    assert tr_title("19 MAYIS") == "19 Mayıs", tr_title("19 MAYIS")
    assert tr_title("GAZİOSMANPAŞA") == "Gaziosmanpaşa", tr_title("GAZİOSMANPAŞA")

    r = parse_row(
        {"OKUL_ADI": "ANKARA - AKYURT - Akyurt Anadolu Lisesi",
         "HOST": "akyurtanadolu", "YOL": "06/26/966556"}, t)
    assert r["il"] == "Ankara", r
    assert r["ilce"] == "Akyurt", r
    assert r["okul_adi"] == "Akyurt Anadolu Lisesi", r
    assert r["kurum_kodu"] == "966556", r
    assert r["okul_turu"] == "Anadolu Lisesi", r
    assert r["web_sitesi"] == "https://akyurtanadolu.meb.k12.tr", r
    assert r["harita"] == "https://akyurtanadolu.meb.k12.tr/tema/harita.php", r
    assert r["adres"] == "" and r["telefon"] == "", r
    assert r["cekim_tarihi"] == t, r

    # Ad içinde ekstra " - " korunur; il/ilçe başlık düzenine çevrilir
    r = parse_row(
        {"OKUL_ADI": "ADANA - SEYHAN - Şehit Ali - Veli İlkokulu",
         "HOST": "x", "YOL": "01/01/123456"}, t)
    assert r["il"] == "Adana" and r["ilce"] == "Seyhan", r
    assert r["okul_adi"] == "Şehit Ali - Veli İlkokulu", r
    assert r["okul_turu"] == "İlkokul", r

    # AFYON -> Afyonkarahisar alias'ı tr_title sonrası uygulanır
    r = parse_row(
        {"OKUL_ADI": "AFYON - MERKEZ - X Lisesi", "HOST": "x",
         "YOL": "03/01/111111"}, t)
    assert r["il"] == "Afyonkarahisar" and r["ilce"] == "Merkez", r

    # Tür çıkarımı — sıra/özgüllük
    cases = {
        "Kadıköy Halk Eğitimi Merkezi": "Halk Eğitimi Merkezi",
        "Çankaya Bilim ve Sanat Merkezi": "Bilim ve Sanat Merkezi",
        "Merkez Şehit X İmam Hatip Ortaokulu": "İmam Hatip Ortaokulu",
        "Y Anadolu İmam Hatip Lisesi": "Anadolu İmam Hatip Lisesi",
        "Z Mesleki ve Teknik Anadolu Lisesi": "Mesleki ve Teknik Anadolu Lisesi",
        "W Çok Programlı Anadolu Lisesi": "Çok Programlı Anadolu Lisesi",
        "Cumhuriyet Ortaokulu": "Ortaokul",
        "Atatürk İlkokulu": "İlkokul",
        "T Fen Lisesi": "Fen Lisesi",
        "Şehit Öğretmen Anaokulu": "Anaokulu",
        "Ahi Evran Mesleki  Eğitim  Merkezi": "Mesleki Eğitim Merkezi",
        "Akşemsettin Ilkokulu": "İlkokul",
        "Adana Adalet Meslekî Eğitim Merkezi": "Mesleki Eğitim Merkezi",
        "Ankara Eğitim ve Uygulama Merkezi": "Özel Eğitim Kurumu",
        "Bir Belediye Nikah Salonu": "Belirsiz",
    }
    for ad, exp in cases.items():
        got = infer_type(ad)
        assert got == exp, f"{ad!r}: beklenen {exp!r}, gelen {got!r}"

    # Bozuk YOL / boş HOST
    r = parse_row({"OKUL_ADI": "A - B - C Lisesi", "HOST": "", "YOL": "06/26"}, t)
    assert r["kurum_kodu"] == "", r
    assert r["web_sitesi"] == "" and r["harita"] == "", r
    assert r["okul_turu"] == "Lise", r

    # Tek parçalı ad
    r = parse_row({"OKUL_ADI": "Garip Kayit", "HOST": "h", "YOL": ""}, t)
    assert r["il"] == "" and r["ilce"] == "" and r["okul_adi"] == "Garip Kayit", r

    # canonicalize: Türkçe karakter içermeyen kopyalar çoğunluğun yazımına eşitlenir
    demo = [
        {"il": "İstanbul", "ilce": "Bağcılar", "okul_adi": "a"},
        {"il": "İstanbul", "ilce": "Bağcılar", "okul_adi": "b"},
        {"il": "İstanbul", "ilce": "Bagcılar", "okul_adi": "c"},
        {"il": "İstanbul", "ilce": "Kadıköy", "okul_adi": "d"},
    ]
    canonicalize(demo)
    assert [x["ilce"] for x in demo] == ["Bağcılar", "Bağcılar", "Bağcılar",
                                        "Kadıköy"], demo

    print("selftest: TÜM TESTLER GEÇTİ")
    return 0


# --- CLI ---------------------------------------------------------------------


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="MEB okul/kurum listesini çeker (CSV/JSON/XLSX/SQLite).")
    p.add_argument("--out-dir", default=".", help="çıktı klasörü (varsayılan: .)")
    p.add_argument("--formats", default="csv,json,xlsx,sqlite",
                   help="virgülle: csv,json,xlsx,sqlite")
    p.add_argument("--il", type=int, default=0,
                   help="sadece bu il kodu (0 = tüm Türkiye; deneme için 81=Düzce)")
    p.add_argument("--delay", type=float, default=0.3,
                   help="sayfalar arası bekleme saniyesi")
    p.add_argument("--selftest", action="store_true",
                   help="ağ kullanmadan ayrıştırma testlerini çalıştır")
    args = p.parse_args(argv)

    if args.selftest:
        return selftest()

    formats = [f.strip() for f in args.formats.split(",") if f.strip()]
    bad = [f for f in formats if f not in WRITERS]
    if bad:
        p.error(f"bilinmeyen format: {', '.join(bad)}")

    return run(args.out_dir, formats, args.il, args.delay)


if __name__ == "__main__":
    raise SystemExit(main())
